from django.contrib import admin
from django.http import HttpResponse
from openpyxl import Workbook
from .models import Persona, Asesor, AsistenteComercial, EjecutivoComercial
from django_pandas.io import read_frame
# Register your models here.

class ExportXLSXMixin:
    def export_as_xlsx(self, request, queryset):

        meta = self.model._meta
        field_names = [field.name for field in meta.fields]

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        file_name = str(meta).replace(".p", "_P")

        response['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(file_name)
        
        workbook  = Workbook()
        worksheet = workbook.active

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(field_names, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title
        df = read_frame(queryset)

        for row_num, row in df.iterrows():
            for col_num, value in enumerate(row):
                cell = worksheet.cell(row=row_num+2, column=col_num+1)
                cell.value = value

        workbook.save(response)
        return response

    export_as_xlsx.short_description = "Exportar XLSX de Personas Seleccionados"

class PersonaAdmin(admin.ModelAdmin, ExportXLSXMixin):

    list_display_links = ('codigo_persona_beme', 'nombre')

    search_fields = ('apellido', 'nombre', 'apellido_materno', 'codigo_persona_beme')

    actions = ["export_as_xlsx"]

    fields = ('codigo_persona_beme', 
                'nombre',
                'apellido',
                'apellido_materno',
                'email',
                'cargo',
                'zona',
                'modulo',
                'sucursal',
                'status')

    list_display = fields #+ ('gerencia', )
    list_filter = ('zona', 'modulo', 'status')


@admin.register(Asesor)
class AsesorAdmin(PersonaAdmin):
    #def formfield_for_foreignkey(self, db_field, request, **kwargs):
    #    if db_field.name == "cargo":
    #        kwargs["queryset"] = Persona.objects.filter(cargo__in=['ASESOR_COMERCIAL'])
    #    return super().formfield_for_foreignkey(db_field, request, **kwargs)
    pass

@admin.register(AsistenteComercial)
class AsistenteComercialAdmin(PersonaAdmin):
    pass

@admin.register(EjecutivoComercial)
class EjecutivoComercialAdmin(PersonaAdmin):

    ordering = ['nombre']

    fields = ('codigo_persona_beme', 
                'nombre',
                'apellido',
                'apellido_materno',
                'email',
                'cargo',
                'zona',
                'modulo',
                'sucursal',
                'status')
    pass

